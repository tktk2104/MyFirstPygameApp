import openpyxl
 
# Excelのファイルを開く
workbook = openpyxl.load_workbook('Project/ExcelFileOperation/res/test.xlsx')
 
# アクティブなシート（先頭のシート）を取得
sheet = workbook.active
 
# セルの値を読み込む(インデックスは１から)
print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=1, column=2).value)
print(sheet.cell(row=1, column=3).value)
print(sheet.cell(row=2, column=1).value)
print(sheet.cell(row=2, column=2).value)
print(sheet.cell(row=2, column=3).value)


## 名前を指定して先頭のシートを取得
#sheet = workbook["Sheet1"]
## インデックスを指定して先頭のシートを取得
#sheet = workbook.worksheets[0]

## 行単位でデータを取得
#for row in sheet.iter_rows(min_row=1):
#    # 行からセルを1個ずつ取得
#    for cell in row:
#    	print(cell.value, end=' ')
#    print() # 行末で改行

#data = [ [cell.value for cell in row  ] for row in sheet]