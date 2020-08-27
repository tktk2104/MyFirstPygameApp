import openpyxl
 
# Excelのワークブックを新規作成
workbook = openpyxl.Workbook()
 
# アクティブなシート（先頭のシート）を取得
sheet = workbook.active
 
# セルに値を書き込む(インデックスは１から)
sheet.cell(row=1, column=1).value = 1
sheet.cell(row=1, column=2).value = 2
sheet.cell(row=1, column=3).value = 3
sheet.cell(row=2, column=1).value = 4
sheet.cell(row=2, column=2).value = 5
sheet.cell(row=2, column=3).value = 6
 
# ファイル名を指定して保存
workbook.save('Project/ExcelFileOperation/res/test.xlsx')